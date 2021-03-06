# Resources:
# Chrome Webdriver: http://stackoverflow.com/questions/8255929/running-webdriver-chrome-with-selenium

import time
import getpass
import psutil
import pyperclip
import sys
import webbrowser
import os
import numpy
import openpyxl
import xlrd
import smtplib
import pandas
import tkinter as tk
import selenium.webdriver.support.ui as ui
import selenium.webdriver as webdriver
from tkinter import *
from tkinter import  ttk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.support import ui
from selenium.webdriver.common.keys import Keys
from pyperclip import copy, paste
from time import sleep
from openpyxl import load_workbook
#from mechanize import Browser
from mechanicalsoup import browser

class App(tk.Frame):
    def __init__(self, master):
        # Grid Layout for Buttons
        # +---------+---------+---------+---------+
        # | (0 , 0) | (0 , 1) | (0 , 2) | (0 , 3) |
        # | (1 , 0) | (1 , 1) | (1 , 2) | (1 , 3) |
        # | (2 , 0) | (2 , 1) | (2 , 2) | (2 , 3) |
        # | (3 , 0) | (3 , 1) | (3 , 2) | (3 , 3) |
        # +---------+---------+---------+---------+

        ttk.Label(master, wraplength=300, text='Automation Tools',
                  background='#006A85',
                  font=('Courier', 15, 'bold')).grid(row=0, column=1, columnspan=3, ipadx=5, ipady=5)  # bg = #005A84

        patientPortal = ttk.Button(master, text="Patient Portal", compound=CENTER, command=self.patientPortal)
        patientPortal.grid(row=1, column=1, padx=5, pady=5, sticky='W')

    # TODO: Log into web page
    def page_is_loaded(driver):
        return driver.find_element_by_tag_name("body") != None

    def patientPortal(self):       
        portal = '<Login WebPage>'
        file = "<File Path>\\551-600dof.xlsx"
        wb = load_workbook(file)
        sheet = wb.active

        # IE
        #ie_driver = webdriver.Ie("C:\Python34\IEDriverServer.exe")
        #ie_driver.get(portal)

        # FireFox
        #firefox_driver = webdriver.Firefox()
        #firefox_driver.get(portal)
        #wait = ui.WebDriverWait(firefox_driver, 10)
        #wait.until(page_is_loaded())
        #username_field = firefox_driver.find_element_by_css_selector("#authUsername")
        #username_field.click()
        #username_field.send_keys("username")
        # FireFox

        # Chrome
        chromeDriver = "<PATH>"
        os.environ["webdriver.chrome.driver"] = chromeDriver
        chrome_driver = webdriver.Chrome(chromeDriver)
        chrome_driver.get(portal)

        time.sleep(3)
        #wait = ui.WebDriverWait(chrome_driver, 10)
        #wait.until(page_is_loaded)

        username_field = chrome_driver.find_element_by_name("authUsername")
        username_field.send_keys("**********")

        password_field = chrome_driver.find_element_by_id("authPassword")
        password_field.send_keys("**********")
        password_field.send_keys(Keys.RETURN)

        time.sleep(2)
        # Chrome

        #TODO: Parse web form
        sheet.columns[1]
        sheet.columns[3]
        sheet.columns[5]
        firstRow = True
        thirdRow = True
        fifthRow = True
        # Find Patient Identifier Field
        for mrnNumber in sheet.columns[1]:
            global firstRow
            if firstRow:
                firstRow = False
                continue
            cellValue = mrnNumber.value
            patient_ID_MRN = chrome_driver.find_element_by_id("id_patient_identifier")
            patient_ID_MRN.send_keys(cellValue)
            #time.sleep(3)
            patient_ID_MRN.send_keys(Keys.RETURN)

            time.sleep(2)

            #TODO: Click Link (href)/workaround 3 tabs?
            usrName = chrome_driver.find_element_by_xpath("/html/body/section/div/div/div/div[2]/div[2]/ul/li/div[1]/h3/a").click()
            time.sleep(3)

            #TODO: Implement Security Questions - Select Patient Postal Code
            secQuestion = chrome_driver.find_element_by_xpath("//select[@id='id_security_question']/option[9]").click()
           
	    #TODO: Error Check if no Postal Code

            #TODO: Select Postal Code from Excel
            #zipDone = True
            for zipCode in sheet.columns[5]:
                global fifthRow
                if fifthRow:
                    fifthRow = False
                    continue
                cellZip = zipCode.value
                _zipCode = chrome_driver.find_element_by_id("id_security_question_answer")
                _zipCode.send_keys(cellZip)

                #TODO: Confirm Security Question Answer
                secQuestionAns = chrome_driver.find_element_by_id("id_confirm_security_question_answer")
                secQuestionAns.send_keys(cellZip)
                #print("ZIP CONFIRMED")
                #if zipDone:
                    #zipDone = False
                    #break
                    #continue
                #continue
                print("ZIPCODE CONFIRMED")

                # TODO: Select Email Address from Excel
                #emailDone = True
                for email in sheet.columns[3]:
                    global thirdRow
                    if thirdRow:
                        thirdRow = False
                        continue
                    emailAddr = email.value
                    recipientEmail = chrome_driver.find_element_by_xpath("//input[@id='id_recipient_email']")
                    recipientEmail.send_keys(emailAddr)

                    #TODO: Confirm Email Address
                    recipientEmailConfirm = chrome_driver.find_element_by_id("id_confirm_recipient_email")
                    recipientEmailConfirm.send_keys(emailAddr)
                    print("EMAIL ADDRESS CONFIRMED")
                    #if emailDone:
                        #emailDone = False
                        #break
                    #continue

                    submit = chrome_driver.find_element_by_id("submit-invitation").click()
                    time.sleep(10)
            print("COMPLETED FIRST ROUND")
            break

            """
            #TODO: Select Email Address from Excel
            emailDone = True
            for email in sheet.columns[3]:
                global thirdRow
                if thirdRow:
                    thirdRow = False
                    continue
                emailAddr = email.value
                recipientEmail = chrome_driver.find_element_by_id("id_recipient_email-help")
                recipientEmail.send_keys(emailAddr)
                print("EMAIL ADDRESS")
                if emailDone:
                    emailDone = False
                    break
                #continue
            print("COMPLETED FIRST ROUND")
            #break

            submit = chrome_driver.find_element_by_id("submit-invitation").click()
            break
                 """
            #usrName = chrome_driver.find_elements_by_link_text('href').click()
            #usrName = chrome_driver.find_element_by_xpath(".//a[/html/body/section/div/div/div/div[2]/div[2]/ul/li/div[1]/h3/a]")
            #usrName.click()

	# Issue with logging into page
        # http: // stackoverflow.com / questions / 21330079 / i - o - exception - and -unable - to - find - element - in -ie - using - selenium - webdriver / 21373224  # 21373224

        # TODO: Click Sign in 

        #time.sleep(7)

        #username_field = ie_driver.find_element_by_id("authUsername")
        #username_field = ie_driver.find_element_by_name("authUsername")
        #username_field.click()
        #username_field.send_keys("<PATH>")

      #  password_field = ie_driver.find_element_by_id("authPassword")
      #  password_field.send_keys("<PATH>")

      #  password_field.send_keys(Keys.RETURN)

"""
    def emailAddr(self):
        file = "<File Path>\\551-600dof.xlsx"
        wb = load_workbook(file)
        sheet = wb.active

        sheet.columns[3]
        thirdRow = True

        chromeDriver = "<PATH>"
        os.environ["webdriver.chrome.driver"] = chromeDriver
        chrome_driver = webdriver.Chrome(chromeDriver)

        # TODO: Select Email Address from Excel
        for email in sheet.columns[3]:
            global thirdRow
            if thirdRow:
                thirdRow = False
                continue
            emailAddr = email.value
            recipientEmail = chrome_driver.find_element_by_id("id_recipient_email-help")
            recipientEmail.send_keys(emailAddr)
            print("EMAIL ADDRESS")

    def zipCode(self):
        file = "<File Path>\\551-600dof.xlsx"
        wb = load_workbook(file)
        sheet = wb.active

        sheet.columns[5]
        fifthRow = True

        chromeDriver = "C:\\Users\\<PATH>"
        os.environ["webdriver.chrome.driver"] = chromeDriver
        chrome_driver = webdriver.Chrome(chromeDriver)

        # TODO: Select Postal Code from Excel
        for zipCode in sheet.columns[5]:
            global fifthRow
            if fifthRow:
                fifthRow = False
                continue
            cellZip = zipCode.value
            _zipCode = chrome_driver.find_element_by_id("id_security_question_answer")
            _zipCode.send_keys(cellZip)

            # TODO: Confirm Security Question Answer
            secQuestionAns = chrome_driver.find_element_by_id("id_confirm_security_question_answer")
            secQuestionAns.send_keys(cellZip)
            print("ZIP CONFIRMED")
"""




"""
        # ----------- START CLOSE IEDriverServer PROCESS ------
        PROCNAME = "IEDriverServer.exe"
        for proc in psutil.process_iter():
            if proc.name() == PROCNAME:
                proc.kill()
"""

def main():

	root = Tk()
	# ----- Windows Details
	root.title("** Login **")
	root.resizable(True,True)
	root.geometry("400x300")
	root.configure(background = '#006A85') #'005A85'

	app = App(root)

	root.mainloop()

if __name__ == "__main__": main()
